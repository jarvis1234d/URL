import sys
from openpyxl import Workbook

txt = open('cod.txt', 'r')
wb = Workbook()
sheet = wb.active
#print(ws['A1'].value)
#cell = sheet.cell(row= 2, column= 1)
#cell.value = 'test'
r = 1
for line in txt:
    lat = line[0: line.index(' ')]
    long = line[line.index(' ')+1: len(line)]
    r +=1#print(f'lat = {lat}, long = {long}' )
    #print(r)
    for c in range(1, 3):
        cell = sheet.cell(row= r, column= c)
        if c ==1:
            cell.value = lat
        elif c ==2:
            cell.value = long



wb.save('Coordinates.xlsx')
